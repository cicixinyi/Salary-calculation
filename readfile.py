import os
import sys
import Employee, EmployeeList
from openpyxl import Workbook, load_workbook
import datetime
import calendar

base = os.path.dirname(os.path.realpath(sys.argv[0]))

def getDate():
        month = int(datetime.date.today().strftime("%m"))
        year = int(datetime.date.today().year)
        day = int(datetime.date.today().day)
        lastDay = calendar.monthrange(year, month)[1]
        # print(day, month, year)
        Date = [year, month, day, lastDay]
        return Date
""" 
def setDate(timelist):
        lastDay = calendar.monthrange(timelist[0],timelist[1])[1]
        timelist.append(lastDay) 
        return timelist """

def readfile(year,month,day):
       # current_m = getDate()[1]
       lastDay = calendar.monthrange(year, month)[1]
       filename=os.path.join(base,"{}工资单.xlsx".format(year))
       list_wb = load_workbook(filename)
       sheet = list_wb["{}月".format(month)]
       # 复制上个月工资表
       if day == lastDay and month < 12:
                target = list_wb.copy_worksheet(sheet)
                target.title = "{}月".format(month+1)
                for row in range(2,target.max_row+1):
                        target.cell(row=row,column=37).value = 0.00   
                        for col in range(4,35):
                            target.cell(row=row,column=col).value = 0.00    
                list_wb.save(filename)
                print('created')
       # 复制第二年工资表
       elif month == 12 and day == lastDay:
              new_wb = Workbook()
              ws = new_wb.create_sheet()
              ws.title = '1月'
              for i,row in enumerate(sheet.iter_rows()):
                for j,cell in enumerate(row):
                        ws.cell(row=i+1, column=j+1, value=cell.value)
              for row in range(2,ws.max_row):
                      ws.cell(row=row,column=37).value = 0.00  
                      for col in range(4,35):
                          ws.cell(row=row,column=col).value = 0.00    
              new_wb.save("/Users/xinyixu/Desktop/工资/{}工资单.xlsx".format(year+1))
              print('created')

       currentSheet =  list_wb[str(month)+"月"]
       employee_info = {}    
       for row in currentSheet.iter_rows(min_row=2, values_only=True):    
                name = row[1] 
                employee_info[name] = {
                        'name': name,
                        'salary': row[2],
                        'totalHour': row[day+1],
                        'totalSalary': row[35],
                        'meal': row[36],
                        'payment': row[37]
                }
       
       # print(employee_info)
       list = EmployeeList.EmployeeList(employee_info) 
       employ_dic = {}
       currentren = list.getHead()
       while currentren is not None:
              employ_dic[currentren.getName()] = currentren
              currentren = currentren.next_employee
       # print(list)
       # print(employ_dic)
       return list, employ_dic

# readfile(2023,5,31)